"""Pipeline node: load candidate-websites from Google Sheets API or local xlsx.

Primary: Google Sheets API (native Google Sheet).
Fallback: local xlsx file via WEBSITES_XLSX_PATH env var.
"""
from __future__ import annotations

import io
import json
import logging
import os
from datetime import datetime, timezone
import re
import unicodedata
from hashlib import sha256
from pathlib import Path
from typing import Any

import aiohttp

from src.services.data_pipeline.base import (
    DataSourceNode,
    NodeConfig,
    register_node,
    save_checkpoint,
)
from src.services.data_pipeline.population import get_top_communes

logger = logging.getLogger(__name__)

SHEETS_API_URL = "https://sheets.googleapis.com/v4/spreadsheets"
SHEETS_SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]

# ---------------------------------------------------------------------------
# Module-level cache so other nodes can access the result
# ---------------------------------------------------------------------------
_cached_websites: dict[tuple[str, str], str] | None = None


def get_websites() -> dict[tuple[str, str], str] | None:
    """Return the websites dict populated by the last run, or ``None``."""
    return _cached_websites


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _norm(s: str) -> str:
    """Normalize a string for fuzzy matching (strip accents, lowercase, alpha only)."""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z]", "", s.lower())


def _get_sheets_credentials():
    """Build Google service-account credentials with Sheets scope."""
    from google.auth.transport.requests import Request
    from google.oauth2.service_account import Credentials

    import base64
    b64 = os.environ.get("GOOGLE_SHEETS_CREDENTIALS_BASE64", "")
    raw = os.environ.get("GOOGLE_SHEETS_CREDENTIALS_JSON", "")
    if b64:
        raw = base64.b64decode(b64).decode()
    elif not raw:
        raise RuntimeError("GOOGLE_SHEETS_CREDENTIALS_JSON env var is not set")
    raw = raw.strip().strip("'\"")
    info = json.loads(raw)
    creds = Credentials.from_service_account_info(info, scopes=SHEETS_SCOPES)
    creds.refresh(Request())
    return creds


# ---------------------------------------------------------------------------
# Node implementation
# ---------------------------------------------------------------------------
class WebsitesNode(DataSourceNode):
    node_id = "websites"
    label = "Candidate Websites"
    default_settings: dict[str, Any] = {
        "sheet_id": "15Mge7CUwsFMn5h7SVRYoo5V1SyDE2vU5h4F9OnDHWB8",
    }

    # --- Loaders -----------------------------------------------------------

    async def _load_sheets_api(
        self, file_id: str, cfg: NodeConfig, force: bool
    ) -> tuple[list[list[str]] | None, str]:
        """Fetch all rows from a native Google Sheet via the Sheets API.

        Returns (rows, content_hash) or (None, hash) if unchanged.
        """
        creds = _get_sheets_credentials()
        headers = {"Authorization": f"Bearer {creds.token}"}

        # Fetch all values from the first sheet
        url = f"{SHEETS_API_URL}/{file_id}/values/A:Z"
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers) as resp:
                resp.raise_for_status()
                data = await resp.json()

        all_rows: list[list[str]] = data.get("values", [])
        if not all_rows:
            raise ValueError("Google Sheet is empty")

        # Content hash for change detection
        raw = json.dumps(all_rows, ensure_ascii=False).encode()
        content_hash = f"sha256:{sha256(raw).hexdigest()}"

        stored_hash = cfg.checkpoints.get("source_hash")
        if not force and stored_hash == content_hash:
            logger.info("[websites] sheet content unchanged, skipping")
            return None, content_hash

        logger.info("[websites] fetched %d rows from Google Sheets", len(all_rows))
        return all_rows, content_hash

    def _load_local_xlsx(self, path: str) -> tuple[list[list[str]], str]:
        """Load rows from a local xlsx file."""
        from openpyxl import load_workbook

        src = Path(path)
        if not src.exists():
            raise FileNotFoundError(f"Websites xlsx not found at {path}")

        logger.info("[websites] loading local file %s", path)
        wb = load_workbook(filename=src, read_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Convert to list[list[str]]
        rows: list[list[str]] = []
        for row in raw_rows:
            rows.append([str(c).strip() if c else "" for c in row])

        content_hash = f"sha256:{sha256(json.dumps(rows, ensure_ascii=False).encode()).hexdigest()}"
        return rows, content_hash

    # --- Main run ----------------------------------------------------------

    async def run(self, cfg: NodeConfig, *, force: bool = False) -> NodeConfig:
        global _cached_websites

        local_path = os.environ.get("WEBSITES_XLSX_PATH", "")
        file_id: str = cfg.settings.get(
            "sheet_id", self.default_settings["sheet_id"]
        )

        if local_path:
            rows, content_hash = self._load_local_xlsx(local_path)
        else:
            result = await self._load_sheets_api(file_id, cfg, force)
            rows, content_hash = result
            if rows is None:
                if _cached_websites is None:
                    # Cache empty (server restart) — force re-fetch
                    rows, content_hash = await self._load_sheets_api(file_id, cfg, True)
                else:
                    return cfg  # unchanged and cache populated

        # ------------------------------------------------------------------
        # Parse rows
        # ------------------------------------------------------------------
        header = [c.strip().lower() for c in rows[0]]
        data_rows = rows[1:]

        def col(name: str) -> int:
            try:
                return header.index(name)
            except ValueError:
                raise ValueError(
                    f"Column '{name}' not found in sheet headers: {header}"
                )

        idx_lastname = col("lastname")
        idx_firstname = col("firstname")
        idx_mun_name = col("municipality_name")
        idx_url = col("website_url")

        # ------------------------------------------------------------------
        # Build commune name -> INSEE code mapping
        # ------------------------------------------------------------------
        top_communes = get_top_communes()
        if not top_communes:
            raise RuntimeError(
                "Population node must run before websites node "
                "(get_top_communes() returned None)"
            )

        communes_by_name: dict[str, str] = {}
        for code, info in top_communes.items():
            communes_by_name[_norm(info["nom"])] = code

        # ------------------------------------------------------------------
        # Match rows to seed data
        # ------------------------------------------------------------------
        result_map: dict[tuple[str, str], str] = {}
        total_rows = len(data_rows)
        skipped_no_url = 0
        skipped_no_commune = 0
        matched = 0

        for row in data_rows:
            # Sheets API returns ragged rows — pad if needed
            url = row[idx_url].strip() if len(row) > idx_url else ""
            if not url or not url.startswith(("http://", "https://")):
                skipped_no_url += 1
                continue

            mun_name = row[idx_mun_name].strip() if len(row) > idx_mun_name else ""
            lastname = row[idx_lastname].strip() if len(row) > idx_lastname else ""
            firstname = row[idx_firstname].strip() if len(row) > idx_firstname else ""

            norm_mun = _norm(mun_name)
            norm_last = _norm(lastname)
            norm_first = _norm(firstname)

            code = communes_by_name.get(norm_mun)
            if not code:
                skipped_no_commune += 1
                continue

            result_map[(code, norm_last)] = url
            result_map[(code, norm_first + norm_last)] = url
            matched += 1

        _cached_websites = result_map

        # ------------------------------------------------------------------
        # Update checkpoints and counts
        # ------------------------------------------------------------------
        cfg.checkpoints["source_hash"] = content_hash
        cfg.checkpoints["cached_at"] = datetime.now(timezone.utc).isoformat()
        await save_checkpoint(cfg.node_id, cfg.checkpoints)

        cfg.counts = {
            "total_rows": total_rows,
            "with_url": total_rows - skipped_no_url,
            "matched_to_seed": matched,
            "skipped_no_url": skipped_no_url,
            "skipped_no_commune": skipped_no_commune,
        }

        logger.info(
            "[websites] parsed %d rows → %d matched, %d no url, %d no commune",
            total_rows,
            matched,
            skipped_no_url,
            skipped_no_commune,
        )

        return cfg


register_node(WebsitesNode())
